# @Time : 2022/3/17 17:33
# @Author : Liang er qian
# @File : excel_Utils.py

import openpyxl


# 将读取Excel文件封装成方法


# 利用反射添加属性
class CaseData():
    def __init__(self, dict_case):
        for i in dict_case.items():
            setattr(self, i[0], i[1])


class ReadExcel:

    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取指定文件，sheet页的所有数据
        wb = openpyxl.load_workbook(file_name)  # 打开传进来的execl文件
        sh = wb[sheet_name]  # 获取指定的sheet页
        allCases = []
        rows = list(sh.rows)  # 获取所有表格
        title = []  # 新建一个列表用来存放表头
        # 遍历第一行数据(表头)
        for cell in rows[0]:  # rows[0] 代表表头的(第一行数据)
            title.append(cell.value)  # 将第一行数据添加到title中

        # 遍历其他行数据，和表头数据进行打包，转成字典，放在列表里面
        for row in rows[1:]:  # 遍历从第二行开始到最后一行
            data = []  # 新建空列表，用来存放每行的数据
            for v in row:  # 遍历
                data.append(v.value)  # 添加到data列表中
            case_zip = dict(zip(title, data))  # 通过聚合打包将 标题和每行数据进行打包
            cd = CaseData(case_zip)  # 利用反射将打包好的数据可以通过key获取value
            allCases.append(cd)  # 添加到allCase中
        return allCases  # 返回

    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        allCase = []
        rows = list(sh.rows)  # 获取所有表格
        title = []  # 创建存放标题的列表
        for cell in rows[0]:  # 循环第一行数据
            tatil.append(cell.value)  # 将循环的第一行数据添加到title列表里
        data = []  # 新建个存放数据的列表
        for row in rows[1:]:  # 从第一行数据开始遍历所有行数据
            for i in row:  # 遍历每一行的数据
                data.append(i.value)  # 添加到data列表中
            case_zip = dict(zip(title, data))  # 通过聚合打包将 标题和每行数据进行打包
            cd = CaseData(case_zip)  # 利用反射将打包好的数据可以通过key获取value
            allCase.append(cd)  # 添加到allCase中
            return allCase  # 返回

    @classmethod
    def wirte_data(cls, file_name, sheet_name, row, column, value):  # 写入指定内容
        """
        :param file_name 文件名
        :param sheet_name sheet页
        :param row 需要写入哪一行
        :param column 需要写入哪一列
        :param value 需要写入的值
        """
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        wb.save(file_name)
